import os
import sys
import csv
import json
import contextlib
from collections import defaultdict

import xlrd
import openpyxl


ID = 'ID'
TITLE = 'TEMA'
LANG = 'IDIOMA'
FORMAT = 'SOPORTE'
ERROR = 'ERROR DE COPIA'
UNKNOWN = 'NO IDENTIFICADO'
REPEATED = 'SE REPITE EN EL INVENTARIO'
REPEATED2 = 'REPETIDO EN EL INVENTARIO'


def translate_header(header):
    translation_table = {
        ID: 'book_id',
        TITLE: 'title',
        LANG: 'lang',
        FORMAT: 'format',
        ERROR: 'error',
        UNKNOWN: 'identified',
        REPEATED: 'repeated',
        REPEATED2: 'repeated'
    }

    return [translation_table.get(h, h) for h in header]


def parse_value(cell):
    if cell.ctype is 1:
        return cell.value.strip()
    elif cell.ctype is 2:
        return int(cell.value) if cell.value.is_integer() else cell.value
    return cell.value


def parse_row(row):
    return [parse_value(cell) for cell in row]


def parse_row_data(header, row):
    if len(row) != len(header):
        raise Exception('Row and header differs in size')
    return dict(zip(header, parse_row(row)))


def transform_values(entry):
    entry['identified'] = not bool(entry.get('identified'))
    entry['error'] = bool(entry.get('error'))
    entry.pop('repeated', None)
    return entry


def parse_rows(sheet):
    header = translate_header(parse_row(sheet.row(0)))

    for line in range(1, sheet.nrows):
        try:
            data = transform_values(parse_row_data(header, sheet.row(line)))
            data['pos'] = line
            yield data
        except Exception as error:
            raise Exception('Error while parsing "{}" line "{}": {}'
                            .format(sheet.name, line, error))


def get_entries_by_id(entries):
    d = defaultdict(list)

    for data in entries:
        d[data['book_id']].append(data)

    return d


def get_diff(old, new):
    old, new = set(old), set(new)

    added = new - old
    removed = old - new
    unchanged = old & new

    return added, removed, unchanged


def open_book(filename):
    book = xlrd.open_workbook(filename)
    pages = book.nsheets
    return book, pages


def parse_listing(filename):
    book = xlrd.open_workbook(filename)
    sheet = book.sheet_by_index(0)
    header = ('book_id', 'key', 'title')
    for n in range(1, sheet.nrows):
        row = sheet.row(n)
        if len(row) != len(header):
            raise Exception('Row {} is not of size {}'.format(n, len(header)))
        yield dict(zip(header, [parse_value(cell) for cell in row]))


def parse_metadata(filename):
    for entry in read(filename):
        entry['id'] = int(entry['id'])
        entry['bb'] = str(entry['bb']).lower() != 'false'
        try:
            entry['ref'] = int(entry['ref'])
        except (ValueError, TypeError):
            pass

        yield entry


def _old_parse_metadata(filename):
    book = xlrd.open_workbook(filename)
    sheet = book.sheet_by_index(0)
    header = [parse_value(cell) for cell in sheet.row(0)]
    for n in range(1, sheet.nrows):
        data = dict(zip(header, (parse_value(cell) for cell in sheet.row(n))))
        for key in ('id', 'lid', 'pos'):
            data[key] = int(data[key])

        for key in ('bb', 'exists'):
            data[key] = str(data[key]).lower() == 'true'

        yield data


class Page(object):
    def __init__(self, n, sheet):
        self.n = n
        self.sheet = sheet
        self.name = sheet.name
        self.nrows = sheet.nrows
        self._entries = None
        self._entries_by_id = None

    @property
    def entries(self):
        if self._entries is None:
            self._entries = list(parse_rows(self.sheet))
        return self._entries

    @property
    def entries_by_id(self):
        if self._entries_by_id is None:
            self._entries_by_id = get_entries_by_id(self.entries)
        return self._entries_by_id

    def diff(self, page):
        return get_diff(self.entries_by_id.keys(), page.entries_by_id.keys())


class Book(object):
    def __init__(self, filename):
        self.filename = filename
        self.book, self.npages = open_book(filename)

    @property
    def pages(self):
        return (self.page(n) for n in range(self.npages))

    def page(self, n):
        return Page(n, self.book.sheet_by_index(n))


_listings = [
    {'year': '1455'},
    {'year': '1553'},
    {'year': '1615'},
    {'year': '1647'},
    {'year': '1726'}
]


def parse_mentions(book):
    for page in book.pages:
        for entry in page.entries:
            yield dict(
                book_id=entry['book_id'],
                list_id=page.n + 1,
                year=_listings[page.n]['year'],
                title=entry['title'],
                key=str(entry['book_id']),
                pos=entry['pos']
            )


def write(data, path=None, **kwargs):
    """Write data to file in any format

    :param data: List of dicts to dump
    :param path: Path of the file to write
    :param header: List of headers to get from data (optional)
    :param format: Name of the format to use (optional)
    """
    ext = format_name(kwargs.pop('format', None) or detect_format(path))
    return dict(
        xlsx=write_xlsx,
        csv=write_csv,
        jsonl=write_jsonl,
    )[ext](data=data, path=path, **kwargs)


def write_xlsx(data, path=None, header=None):
    header = resume_header(data) if header is None else header
    book = openpyxl.Workbook()
    sheet = book.active

    for n, key in enumerate(header, 1):
        sheet.cell(column=n, row=1, value=key)

    for row, entry in enumerate(data, 2):
        for col, key in enumerate(header, 1):
            sheet.cell(column=col, row=row, value=entry.get(key))

    if path:
        book.save(path)

    return book


def write_csv(data, path, header=None):
    header = resume_header(data) if header is None else header
    with open_file(path, 'w') as stream:
        writer = csv.writer(stream)
        writer.writerow(header)
        for entry in data:
            writer.writerow(entry.get(k) for k in header)


def write_jsonl(data, path=None, header=None):
    if header is None:
        header = resume_header(data)
    with open_file(path, 'w') as stream:
        line = json.dumps(
            [{k: entry.get(k) for k in header} for entry in data]
        )
        stream.write(line + '\n')


FORMATS = ('xlsx', 'csv', 'jsonl')
FORMAT_ALIASES = {
    'xls': 'xlsx',
    'jl': 'jsonl'
}


@contextlib.contextmanager
def open_file(path, *args, **kwargs):
    if not path or path == '-':
        yield sys.stdout
    else:
        with contextlib.closing(open(path, *args, **kwargs)) as stream:
            yield stream


def format_name(ext, default='csv'):
    ext = FORMAT_ALIASES.get(ext, ext or default)
    if ext not in FORMATS:
        raise ValueError('Unknown format "{}" should be one of {}'
                         .format(ext, ', '.join(FORMATS)))
    return ext


def detect_format(path):
    return format_name(path and os.path.splitext(path)[-1].strip('.'))


def read(path, **kwargs):
    """Read data from given format

    :param path: Path from where to read the data
    :param header: List of fields to save (optional)
    :param types: key: function dict for type conversions
    :returns: Sequence of dicts with the fields
    """
    ext = format_name(kwargs.pop('format', None) or detect_format(path))
    return dict(
        xlsx=read_xlsx,
        csv=read_csv,
        jsonl=read_jsonl,
    )[ext](path, **kwargs)


def resume_header(sequence):
    return sorted(set(k for e in sequence for k in e))


def sequence_to_listdict(sequence, header, types=None):
    types = {} if types is None else types
    for values in sequence:
        entry = dict(zip(header, values))
        for name, function in types.items():
            entry[name] = function(entry[name])
        yield entry


def read_xlsx(path, header=None, **kwargs):
    rows = openpyxl.load_workbook(path).active.rows
    if header is None:
        header = list(c.value for c in next(rows) if c.value)
    non_empty_rows = (
        (c.value for c in row) for row in rows if any((c.value for c in row))
    )
    yield from sequence_to_listdict(non_empty_rows, header, **kwargs)


def read_csv(path, header=None, **kwargs):
    with open(path) as stream:
        rows = csv.reader(stream)
        inline_header = next(rows)
        if header is None:
            header = inline_header
        yield from sequence_to_listdict(rows, header, **kwargs)


def read_jsonl(path, header=None, **kwargs):
    if header is None:
        with open(path) as stream:
            header = resume_header(json.loads(line) for line in stream)

    with open(path) as stream:
        rows = ((data.get(h) for h in header)
                for data in (json.loads(line) for line in stream))
        yield from sequence_to_listdict(rows, header, **kwargs)


def merge(left, right, key, right_key=None, header=None,
          left_header=None, right_header=None, only_left=None):
    left_key = key
    if right_key is None:
        right_key = left_key
    if left_header is None:
        left = list(left)
        left_header = resume_header(left)
    if right_header is None:
        right = list(right)
        right_header = resume_header(right)
    if header is None:
        header = set(left_header) | set(right_header)

    indexed = {entry[right_key]: entry for entry in right}
    for lentry in left:
        left_key_value = lentry[left_key]
        rentry = indexed.get(left_key_value, {})

        entry = {k: lentry.get(k) for k in left_header}
        entry.update(
            (k, rentry.get(k, rentry.get(k))) for k in right_header
        )

        yield {k: entry[k] for k in header}
